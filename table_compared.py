import openpyxl
import pandas as pd
import time


def timing_decorator(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        execution_time = end_time - start_time
        print(f"Время выполнения функции {func.__name__}: {execution_time:.6f} секунд")
        return result
    return wrapper


def excel_to_dict(file_path_):
    try:
        df = pd.read_excel(file_path_)
        data_dict_ = {}

        for col in df.columns:
            data_dict_[col] = df[col].tolist()

        return data_dict_
    except FileNotFoundError:
        print('Одного из сравниваемых файлов не существует!')


def export_excel(worksheet, value, column, row):
    worksheet.cell(row=row, column=column, value=value)


@timing_decorator
def compared(original, compared_, percent):
    print('Функция "compared()" сравнивает данные таблиц ...')
    data_dict = excel_to_dict(original)
    article_list = data_dict['Артикул, id']
    price_list = data_dict['Цена со скидкой']
    name_list = data_dict['Наименование']

    data_dict2 = excel_to_dict(compared_)
    article_list2 = data_dict2['Артикул, id']
    price_list2 = data_dict2['Цена со скидкой']

    workbook = openpyxl.load_workbook(original)
    sheet = workbook.active

    list_dumping = []
    for idx_2, id_2 in enumerate(article_list2):
        if id_2 in article_list:
            idx_ = article_list.index(id_2)
            try:
                difference = (price_list[idx_] - price_list2[idx_2]) / price_list[idx_] * 100
            except ZeroDivisionError:
                difference = 0
            if difference > percent:
                                #   [название товара,  id,   % изменения цены,       первая цена,      текущая цена  ]
                list_dumping.append([name_list[idx_], id_2, round(difference, 2), price_list[idx_], price_list2[idx_2]])
            export_excel(sheet, price_list2[idx_2], 5, idx_ + 2)
            export_excel(sheet, difference, 6, idx_ + 2)
            export_excel(sheet, id_2, 7, idx_ + 2)

    workbook.save(original)

    return list_dumping


def compared_vi(original):
    workbook = openpyxl.load_workbook(original)
    sheet = workbook.active
