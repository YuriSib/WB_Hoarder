import openpyxl
import time
from collections import Counter

from search_sale import ya_search


def timing_decorator(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        execution_time = end_time - start_time
        print(f"Время выполнения функции {func.__name__}: {execution_time:.6f} секунд")
        return result
    return wrapper


@timing_decorator
def link_collecting(table):
    workbook = openpyxl.load_workbook(table)
    sheet = workbook.active
    list_link_ = []
    for row in range(2, sheet.max_row-1):
        cell_value = sheet.cell(row=row, column=2).value
        links, list_link_vi, list_link_ozon = ya_search(cell_value, 20)
        for link in links:
            if 'https://market.yandex.ru' not in link and 'https://www.vseinstrumenti.ru' not in link \
                    and 'https://www.WildBerries.ru' not in link:
                list_link_.append(link)

    return list_link_


def link_separation(list_):
    list_link_ = []
    for link in list_:
        start_index = link.find('https://') + len('https://')
        end_index = link.find('/', start_index)
        result = link[start_index:end_index]
        list_link_.append(result)

    count_elements = Counter(list_link_)
    sorted_elements = sorted(count_elements.items(), key=lambda x: x[1], reverse=True)

    for element, count in sorted_elements:
        print(f"{element}: {count}")


test_table = 'compared — копия.xlsx'
list_link = link_collecting(test_table)
link_separation(list_link)

