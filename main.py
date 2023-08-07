from openpyxl import Workbook, load_workbook
import openpyxl
import time

from collecting import hoarder
from openpyxl.utils.dataframe import dataframe_to_rows


def timing(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        execution_time = end_time - start_time
        print(f"Время выполнения функции {func.__name__}: {execution_time:.6f} секунд")
        return result
    return wrapper


def save_in_excel(data_list, path):
    try:
        workbook = load_workbook(path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
    headers = list(data_list[0].keys())

    if sheet.max_row == 1:
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)

    next_row = sheet.max_row + 1

    for item in data_list:
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=next_row, column=col_idx, value=item[header])
        next_row += 1

    workbook.save(path)


def changes_prise_in_table(path, row, column, value):
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.cell(row=row, column=column, value=value)
    workbook.save(path)


def find_value_row(file_path, value):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell.row

    return None


@timing
def pars_and_save(table):
    print('Функция "pars_and_save()" сканирует категерию ВБ и сохраняет данные о товарах в таблицу ...')
    specification_list = []
    count = 1
    while True:
        if hoarder(count):
            specification = hoarder(count)
            save_in_excel(specification, table)
            specification_list.append(specification)
            count += 1
        else:
            break
    return specification_list


# original_table = 'original.xlsx'
# compared_table = 'compared.xlsx'
#
# changes_prise_in_table(compared_table, 1, 2, 'Prise was changes!')
